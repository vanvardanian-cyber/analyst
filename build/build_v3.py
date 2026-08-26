#!/usr/bin/env python3
"""Build v3: full Amazon.de seller workbook — navigation, guidance, all modules."""
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter

OUT = "/home/claude/amazon-de-seller-workbook-v3-EN.xlsx"
wb = Workbook()

# ---------- styles ----------
F_TITLE   = Font(name="Arial", size=12, bold=True)
F_SECTION = Font(name="Arial", size=10, bold=True)
F_BASE    = Font(name="Arial", size=10)
F_INPUT   = Font(name="Arial", size=10, color="0000FF")
F_FORMULA = Font(name="Arial", size=10)
F_LINK    = Font(name="Arial", size=10, color="008000")
F_NOTE    = Font(name="Arial", size=9, italic=True, color="808080")
F_BOLD    = Font(name="Arial", size=10, bold=True)
F_NAV     = Font(name="Arial", size=10, color="0563C1", underline="single")
FILL_IN   = PatternFill("solid", fgColor="FFFF00")
FILL_HDR  = PatternFill("solid", fgColor="D9D9D9")
WRAP      = Alignment(wrap_text=True, vertical="top")

NUM, EUR, PCT, IDX, DATEF = "#,##0", '#,##0.00" €"', "0.0%", "0.00", "yyyy-mm-dd"

TAB = {"nav": "808080", "choose": "4472C4", "launch": "70AD47", "operate": "ED7D31"}

def put(ws, coord, value, font=F_BASE, fill=None, fmt=None, wrap=False):
    c = ws[coord]; c.value = value; c.font = font
    if fill: c.fill = fill
    if fmt:  c.number_format = fmt
    if wrap: c.alignment = WRAP
    return c

def inp(ws, coord, value, fmt=None):  return put(ws, coord, value, font=F_INPUT, fill=FILL_IN, fmt=fmt)
def frm(ws, coord, value, fmt=None, font=F_FORMULA): return put(ws, coord, value, font=font, fmt=fmt)
def note(ws, coord, value): return put(ws, coord, value, font=F_NOTE, wrap=True)
def widths(ws, spec):
    for col, w in spec.items(): ws.column_dimensions[col].width = w
def backlink(ws):
    frm(ws, "B1", '=HYPERLINK("#Start!B2","← Start")', font=F_NAV)
def comment(ws, coord, text):
    ws[coord].comment = Comment(text, "Workbook", height=140, width=280)

# =====================================================================
# START (dashboard / navigation)
# =====================================================================
ws = wb.active; ws.title = "Start"; ws.sheet_properties.tabColor = TAB["nav"]
widths(ws, {"A": 2, "B": 20, "C": 46, "D": 9, "E": 48})
put(ws, "B2", "Amazon.de Seller Workbook — v3", F_TITLE)
put(ws, "B3", "Choose the product, plan the order, track the money — one file. Follow the order below for a new product; every sheet links back here.", F_BASE, wrap=True)
put(ws, "B5", "Legend: BLUE on YELLOW = you type here · BLACK = formula, do not overwrite · GREEN = pulled from another sheet automatically.", F_NOTE, wrap=True)
put(ws, "B6", "Cells with a small red corner marker have an (i) note — hover to read what the number means and what to do with it.", F_NOTE, wrap=True)

hdr_r = 8
for col, h in [("B","Sheet"),("C","What it answers"),("D","Open"),("E","Where you stand (live)")]:
    put(ws, f"{col}{hdr_r}", h, F_SECTION, fill=FILL_HDR)

nav = [
 ("Guide", "How to use the file + where every report lives", '"read once before starting"'),
 ("1 Seasonality", "Is demand seasonal? When does it ramp?",
  "IFERROR(\"amplitude \"&TEXT('1 Seasonality'!C29,\"0.00\")&\"x · ramp \"&'1 Seasonality'!C34&\" · YoY \"&TEXT('1 Seasonality'!C31,\"+0.0%;-0.0%\"),\"fill Sheet 1\")"),
 ("2 Order timing", "By when must the order be placed?",
  "IFERROR(\"order by \"&TEXT('2 Order timing'!C20,\"yyyy-mm-dd\")&\" · \"&TEXT('2 Order timing'!C21,\"0\")&\" days left\",\"fill Sheet 2\")"),
 ("3 Unit economics", "Do I earn enough per unit before ads?",
  "IFERROR(\"margin \"&TEXT('3 Unit economics'!C30,\"0.0%\")&\" · breakeven ACOS \"&TEXT('3 Unit economics'!C31,\"0.0%\"),\"fill Sheet 3\")"),
 ("4 Selection", "Does the product pass every gate?",
  "IFERROR(\"A: \"&'4 Selection'!C32&\" · B \"&'4 Selection'!C33&\" · C \"&'4 Selection'!C34,\"fill Sheet 4\")"),
 ("5 Niche scorecard", "Which candidate deserves my research time?",
  "IF(COUNT('5 Niche scorecard'!D32:K32)=0,\"no niches scored yet\",\"best score \"&TEXT(MAX('5 Niche scorecard'!D32:K32),\"0\")&\"/100\")"),
 ("6 Order size", "How many units in the first order?",
  "IFERROR(\"order \"&TEXT('6 Order size'!C24,\"#,##0\")&\" units\",\"fill Sheets 3 + 6\")"),
 ("7 PPC planner", "What can I pay for ads? Where is waste?",
  "IFERROR(\"target ACOS \"&TEXT('7 PPC planner'!C11,\"0.0%\")&\" · wasted \"&TEXT('7 PPC planner'!C25,\"#,##0\")&\" € found\",\"fill Sheet 7\")"),
 ("8 Cash flow", "Do I stay above zero the whole year?",
  "IFERROR(\"lowest cash point \"&TEXT('8 Cash flow'!C32,\"#,##0\")&\" €\",\"fill Sheet 8\")"),
 ("9 Monthly P&L", "Am I actually making money each month?",
  "IF('9 Monthly P&L'!K18=0,\"no months filled yet\",\"YTD profit \"&TEXT('9 Monthly P&L'!K18,\"#,##0\")&\" €\")"),
 ("10 Money recovery", "Does Amazon owe me money?",
  "\"unclaimed flags: \"&'10 Money recovery'!C5"),
 ("11 Review themes", "Which complaint do I attack? Is the wall fake?",
  "IF('11 Review themes'!C19=\"—\",\"no themes entered yet\",'11 Review themes'!C19&\" (\"&TEXT('11 Review themes'!C20,\"0%\")&\")\")"),
]
for i, (name, what, live) in enumerate(nav):
    r = hdr_r + 1 + i
    put(ws, f"B{r}", name, F_BOLD)
    put(ws, f"C{r}", what, F_BASE, wrap=True)
    frm(ws, f"D{r}", f'=HYPERLINK("#\'{name}\'!B2","Open →")', font=F_NAV)
    frm(ws, f"E{r}", f"={live}", font=F_LINK)

put(ws, "B22", "New product path:  1 → 2 → 3 → 4  (run 5 to compare several niches)  → 6 → 7 → 8, then order.", F_SECTION)
put(ws, "B23", "Already selling:  9 → 10 → 11 once a month, 7 every two weeks.", F_SECTION)
note(ws, "B25", "v3 · 2026-08-18 · Built for a non-EU (deemed supplier) seller on Amazon.de. Educational tool, not tax or legal advice.")

# =====================================================================
# GUIDE
# =====================================================================
ws = wb.create_sheet("Guide"); ws.sheet_properties.tabColor = TAB["nav"]
widths(ws, {"A": 2, "B": 110}); backlink(ws)
put(ws, "B2", "Guide: how to use this workbook", F_TITLE)
guide_rows = [
 (4,  "The rules", F_SECTION),
 (5,  "Blue text on yellow fill = you type here. Black = formula, never overwrite. Green = pulled from another sheet. Hover cells with a red corner marker for (i) explanations. Work sheets 1→8 in order for a new product; sheets 9–11 are your monthly routine after launch.", F_BASE),
 (7,  "Measure seasonality BEFORE profit math — it is faster and kills more candidates. Never measure seasonality on a listing younger than 18 months: a young listing's chart is launch curve + season, inseparable. Measure the CATEGORY on data that existed before the listing did.", F_BASE),
 (9,  "Where the research data comes from (sheets 1–5)", F_SECTION),
 (10, "Search history: Helium 10 Cerebro/Magnet → chart icon next to search volume → export. Amazon's own data: Product Opportunity Explorer in Seller Central (free). Old competitors' sales: Helium 10 Trendster, 3–5 ASINs older than 2 years. Same-month-last-year: Xray columns 'Last Year Sales' / 'Sales YoY'. Shape check: Google Trends, region Germany, 5-year window (relative values only — shape, never volume).", F_BASE),
 (13, "BSR is a rank, weighted to recent sales: if a whole category halves in December, ranks barely move, so BSR UNDERSTATES seasonal amplitude. Use units or search volume for amplitude; BSR only to compare products at one moment.", F_BASE),
 (15, "Where the operating reports live (sheets 7, 9, 10)", F_SECTION),
 (16, "Search Term Report (Sheet 7): Advertising console → Sponsored Products → Reports → Search term. Choose 30–60 days.", F_BASE),
 (17, "Date Range / transaction totals (Sheet 9): Seller Central → Payments → Reports repository → Date Range Report (Summary).", F_BASE),
 (18, "Inventory Adjustments (Sheet 10): Reports → Fulfilment by Amazon → Inventory → Inventory Adjustments. Filter reasons: Lost, Damaged, Disposed.", F_BASE),
 (19, "Reimbursements (Sheet 10): Reports → Fulfilment by Amazon → Payments → Reimbursements.", F_BASE),
 (20, "Reviews (Sheet 11): Helium 10 → Review Insights → export 1–3★ reviews of the top 5–10 ASINs, last 12 months. Cluster into themes with an AI assistant ('cluster these reviews into complaint themes with counts'), then enter the themes.", F_BASE),
 (22, "Method sources", F_SECTION),
 (23, "Seasonal decomposition: Hyndman & Athanasopoulos, Forecasting: Principles and Practice (free: https://otexts.com/fpp3/). Order sizing: the newsvendor model (operations research standard). Review mining: Archak, Ghose & Ipeirotis, Management Science 2011. Fake-review signals: He, Hollenbeck & Proserpio, Marketing Science 2022.", F_BASE),
 (25, "What changed in v3", F_SECTION),
 (26, "Added: Start dashboard with live status per sheet · Sheet 6 first-order size (newsvendor) · Sheet 7 PPC planner with breakeven/target ACOS, TACOS, ROAS and a search-term waste finder · Sheet 8 twelve-month cash-flow plan incl. import-VAT refund timing · Sheet 9 monthly P&L with TACOS tracking · Sheet 10 money-recovery audit (lost inventory matcher + FBA fee sanity check) · Sheet 11 review-theme analysis + fake-review checklist · (i) comments and back-links everywhere.", F_BASE),
 (28, "Carried over from v2 (fixes to the original file): duty on CIF base · breakeven ACOS vs VAT-inclusive price · max EXW supplier ceiling · Georgian 1% turnover tax · Dec→Jan ramp wrap · incomplete-data guard on the seasonality verdict · IP + PPWR rows in the selection gates.", F_BASE),
 (30, "This file is an educational tool, not tax, legal or investment advice. Amazon fees change every December/January — cross-check your category's rates in Seller Central before big decisions.", F_NOTE),
]
for r, text, font in guide_rows:
    put(ws, f"B{r}", text, font, wrap=True)

# =====================================================================
# 1 SEASONALITY (v2 sheet + backlink + comments)
# =====================================================================
ws = wb.create_sheet("1 Seasonality"); ws.sheet_properties.tabColor = TAB["choose"]
widths(ws, {"A": 2, "B": 26, "C": 12, "D": 12, "E": 12, "F": 9, "G": 9, "H": 30, "I": 6})
backlink(ws)
put(ws, "B2", "Seasonality index from monthly demand", F_TITLE)
put(ws, "B3", "Enter 24 months of search volume for the niche keyword. The sheet computes the shape of the year and a verdict.", F_NOTE, wrap=True)
put(ws, "B4", "Keyword / niche", F_SECTION);  inp(ws, "C4", "fahrradhelm")
note(ws, "G4", "Example shown. Replace with your own export.")
put(ws, "B5", "Marketplace", F_SECTION);      inp(ws, "C5", "Amazon.de")
put(ws, "B6", "Data source", F_SECTION);      inp(ws, "C6", "H10 Cerebro Search Volume History")
put(ws, "B7", "Export date", F_SECTION);      inp(ws, "C7", datetime.date(2026, 8, 14), fmt=DATEF)

for j, h in enumerate(["Month","Volume yr 1","Volume yr 2","Avg volume","Index","YoY %","Notes","Flag"]):
    put(ws, f"{get_column_letter(2+j)}9", h, F_SECTION, fill=FILL_HDR)
months = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
y1 = [9800,11500,21000,29000,34000,38000,41000,33000,22000,14000,9200,8100]
y2 = [10400,12600,23500,31500,37000,41000,45000,36000,24000,15200,9900,8600]
for i, m in enumerate(months):
    r = 10 + i
    put(ws, f"B{r}", m, F_BASE)
    inp(ws, f"C{r}", y1[i], fmt=NUM); inp(ws, f"D{r}", y2[i], fmt=NUM)
    frm(ws, f"E{r}", f'=IF(COUNT(C{r}:D{r})=0,"",AVERAGE(C{r}:D{r}))', fmt=NUM)
    frm(ws, f"F{r}", f'=IF($E{r}="","",$E{r}/AVERAGE($E$10:$E$21))', fmt=IDX)
    frm(ws, f"G{r}", f'=IF(COUNT(C{r}:D{r})<2,"",D{r}/C{r}-1)', fmt=PCT)
    prev = 21 if r == 10 else r - 1
    frm(ws, f"I{r}", f'=IF(OR($F{r}="",$F${prev}=""),0,IF(AND($F{r}>=1,$F${prev}<1),1,0))', font=F_NOTE)
put(ws, "B22", "Total", F_BOLD)
frm(ws, "C22", "=SUM(C10:C21)", fmt=NUM, font=F_BOLD); frm(ws, "D22", "=SUM(D10:D21)", fmt=NUM, font=F_BOLD)
frm(ws, "E22", "=SUM(E10:E21)", fmt=NUM, font=F_BOLD)
frm(ws, "G22", '=IF(OR(C22=0,D22=0),"",D22/C22-1)', fmt=PCT, font=F_BOLD)
put(ws, "B24", "Results", F_SECTION, fill=FILL_HDR); put(ws, "C24", "", F_SECTION, fill=FILL_HDR)
put(ws, "B25", "Peak month", F_BASE)
frm(ws, "C25", '=IF(COUNT($F$10:$F$21)=0,"",INDEX($B$10:$B$21,MATCH(MAX($F$10:$F$21),$F$10:$F$21,0)))')
put(ws, "B26", "Lowest month", F_BASE)
frm(ws, "C26", '=IF(COUNT($F$10:$F$21)=0,"",INDEX($B$10:$B$21,MATCH(MIN($F$10:$F$21),$F$10:$F$21,0)))')
put(ws, "B27", "Peak index", F_BASE);  frm(ws, "C27", '=IF(COUNT($F$10:$F$21)=0,"",MAX($F$10:$F$21))', fmt=IDX)
put(ws, "B28", "Lowest index", F_BASE); frm(ws, "C28", '=IF(COUNT($F$10:$F$21)=0,"",MIN($F$10:$F$21))', fmt=IDX)
put(ws, "B29", "Amplitude (peak / lowest non-zero)", F_BASE)
frm(ws, "C29", '=IFERROR(MAX($F$10:$F$21)/_xlfn.MINIFS($F$10:$F$21,$F$10:$F$21,">0"),"")', fmt=IDX, font=F_BOLD)
comment(ws, "C29", "How many times bigger the best month is vs the worst. <1.5 flat · 1.5-2.5 moderate · 2.5-4 strong · >4 single-season (capital trap for a first product).")
put(ws, "B30", "Top-4 months' share of the year", F_BASE)
frm(ws, "C30", '=IFERROR((LARGE($E$10:$E$21,1)+LARGE($E$10:$E$21,2)+LARGE($E$10:$E$21,3)+LARGE($E$10:$E$21,4))/SUM($E$10:$E$21),"")', fmt=PCT)
put(ws, "B31", "Category growth year over year", F_BASE)
frm(ws, "C31", '=IF(OR(C22=0,D22=0),"",D22/C22-1)', fmt=PCT)
put(ws, "B32", "Months with data (needs 12)", F_BASE); frm(ws, "C32", "=COUNT($E$10:$E$21)", fmt="0")
put(ws, "B33", "Verdict", F_SECTION)
frm(ws, "C33",
    '=IF(C32<12,"Fill all 12 months before trusting the verdict.",'
    'IF(COUNTIF($E$10:$E$21,0)>0,"A zero-demand month: extreme single-season. Avoid as a first product.",'
    'IF(C29<1.5,"Flat demand. Order timing is not critical.",'
    'IF(C29<2.5,"Moderate seasonality. Time the order, but sales run all year.",'
    'IF(C29<4,"Strong seasonality. Stock must land 4 weeks before the ramp; exit the peak with low inventory.",'
    '"Single-season product. A working-capital trap for a first product.")))))', font=F_BOLD)
put(ws, "B34", "Ramp month (index crosses 1.00 upward)", F_BASE)
frm(ws, "C34", '=IFERROR(INDEX($B$10:$B$21,MATCH(1,$I$10:$I$21,0)),"—")', font=F_BOLD)
comment(ws, "C34", "First calendar month where the index crosses 1.00 upward. Dec-Jan wrap handled via flags in column I. If demand is bimodal, check all flags.")
put(ws, "B35", "Demand direction (from YoY)", F_SECTION)
frm(ws, "C35",
    '=IF(C31="","Fill both years first.",'
    'IF(C31>0.05,"GROWING (>+5%). Tailwind. Check it is not one viral ASIN: in Xray, is page-1 revenue spread across many listings?",'
    'IF(C31>=-0.05,"FLAT (-5%..+5%). Fine for entry - you win by taking share, not riding the tide.",'
    'IF(C31>=-0.15,"COOLING (-5%..-15%). Google Trends, DE, 5 years: post-boom plateau = acceptable; steady multi-year slide = drop.",'
    '"DECLINING (<-15%). Fails the demand test - drop unless Trends shows a clearly temporary dip."))))', font=F_BOLD)
comment(ws, "C35", "Two years of data can be distorted by one-off spikes (COVID-type booms, a viral video). Whatever this verdict says, confirm the direction on the 5-year Google Trends curve before deciding.")
note(ws, "B36", "Amplitude = shape of the year; YoY = does the category grow. They answer different questions.")
note(ws, "B37", "Top-4 share above 60% = a four-month business carrying twelve months of costs.")
note(ws, "B38", "YoY verdict bands: >+5% growing · -5..+5 flat · -5..-15 cooling (verify on Trends) · below -15% declining.")

# =====================================================================
# 2 ORDER TIMING
# =====================================================================
ws = wb.create_sheet("2 Order timing"); ws.sheet_properties.tabColor = TAB["choose"]
widths(ws, {"A": 2, "B": 52, "C": 14, "D": 2, "E": 62}); backlink(ws)
put(ws, "B2", "Back-calculating the order date from the ramp month", F_TITLE)
put(ws, "B3", "Target on-stock date = 4 weeks before the ramp month from Sheet 1.", F_BASE)
put(ws, "B5", "Ramp month (from Sheet 1)", F_BASE); frm(ws, "C5", "='1 Seasonality'!C34", font=F_LINK)
put(ws, "B6", "Suggested on-stock date (1st of ramp month − 28 days)", F_BASE)
frm(ws, "C6",
    "=IFERROR(DATE(YEAR(TODAY())+IF(DATE(YEAR(TODAY()),MATCH($C$5,'1 Seasonality'!$B$10:$B$21,0),1)-28<=TODAY(),1,0),"
    "MATCH($C$5,'1 Seasonality'!$B$10:$B$21,0),1)-28,\"fill Sheet 1 first\")", fmt=DATEF)
put(ws, "B7", "Target on-stock date (your choice)", F_BASE); inp(ws, "C7", datetime.date(2027, 3, 4), fmt=DATEF)
note(ws, "E7", "Defaults to the suggested date. Land earlier if peak-season check-in delays are likely.")
put(ws, "B9", "Supply-cycle stage", F_SECTION, fill=FILL_HDR); put(ws, "C9", "Days", F_SECTION, fill=FILL_HDR)
stages = [
 (10, "Certification / test protocols (0 if unregulated)", 0,  "PPE, electrical, toys: add 60–120 days."),
 (11, "Supplier selection and samples", 21, "Two sample rounds is normal."),
 (12, "Production", 40, "30–45 days for a 300–1,000-unit batch."),
 (13, "Factory holiday buffer (CNY, Golden Week)", 0, "Chinese New Year (late Jan–Feb): add 20–30 days. Golden Week (early Oct): add ~7."),
 (14, "Sea freight to Germany", 40, "35–45 days port-to-port plus the inland leg."),
 (15, "Customs clearance", 5, "3–7 days. Non-EU importer: you need an EORI and normally an indirect customs representative (usually the forwarder). Agree incoterms before production starts."),
 (16, "FBA inbound and check-in", 10, "7–14 days, longer in Q4."),
 (17, "Safety buffer", 15, "Never set this to zero."),
]
for r, label, days, n in stages:
    put(ws, f"B{r}", label, F_BASE, wrap=True); inp(ws, f"C{r}", days, fmt="0"); note(ws, f"E{r}", n)
put(ws, "B18", "Total lead time (days)", F_BOLD); frm(ws, "C18", "=SUM(C10:C17)", fmt="0", font=F_BOLD)
put(ws, "B20", "Order must be placed by", F_BOLD); frm(ws, "C20", "=C7-C18", fmt=DATEF, font=F_BOLD)
put(ws, "B21", "Days from today", F_BOLD); frm(ws, "C21", "=C7-C18-TODAY()", fmt="0", font=F_BOLD)
note(ws, "B23", "Negative = the season is already lost. Plan next year or change the product.")

# =====================================================================
# 3 UNIT ECONOMICS
# =====================================================================
ws = wb.create_sheet("3 Unit economics"); ws.sheet_properties.tabColor = TAB["choose"]
widths(ws, {"A": 2, "B": 46, "C": 13, "D": 2, "E": 66}); backlink(ws)
put(ws, "B2", "Contribution margin per unit, Amazon.de", F_TITLE)
put(ws, "B3", "Non-EU seller: Amazon acts as deemed supplier, so the 19% German VAT never reaches you. Net revenue = gross / 1.19.", F_BASE, wrap=True)
put(ws, "B5", "Inputs", F_SECTION, fill=FILL_HDR); put(ws, "C5", "", F_SECTION, fill=FILL_HDR)
inputs3 = [
 (6,  "Sale price, gross incl. VAT (EUR)", 59.99, EUR, "What the buyer pays."),
 (7,  "VAT rate", 0.19, PCT, "German standard rate."),
 (8,  "Referral fee rate", 0.15, PCT, "Check your category — Dec 2025/Jan 2026 updates cut several sub-€20 categories."),
 (9,  "Referral charged on gross price? 1=yes 0=no", 1, "0", "Amazon EU charges referral fees on the VAT-inclusive total sales price — leave at 1."),
 (10, "FBA fulfilment fee (EUR)", 4.55, EUR, "From the listing or the Revenue Calculator."),
 (11, "Storage per unit sold (EUR)", 0.62, EUR, "Litres × monthly rate × months of cover. Oct–Dec rates are ~2–3× higher."),
 (12, "Returns rate", 0.07, PCT, "Size/fit products run 8–12%."),
 (13, "Return cost, share of net price", 0.70, PCT, "Lost fees + processing + unsellable stock."),
 (14, "EXW unit cost (EUR)", 10.0, EUR, "Supplier price."),
 (15, "Freight per unit (EUR)", 2.20, EUR, "Rate per m³ × unit volume."),
 (16, "Import duty rate (on CIF value)", 0.027, PCT, "Check the TARIC code. Duty base = goods + freight, not EXW alone."),
 (17, "Prep, labels, inspection per unit (EUR)", 0.40, EUR, "Incl. GPSR & EPR marking."),
 (18, "Turnover tax rate (home country)", 0.01, PCT, "Georgian small-business IE: 1% of turnover. Confirm the base with your accountant."),
]
for r, label, v, fmt, n in inputs3:
    put(ws, f"B{r}", label, F_BASE); inp(ws, f"C{r}", v, fmt=fmt); note(ws, f"E{r}", n)
put(ws, "B20", "Calculation", F_SECTION, fill=FILL_HDR); put(ws, "C20", "", F_SECTION, fill=FILL_HDR)
calc3 = [
 (21, "Net revenue (ex VAT)", "=C6/(1+C7)", EUR, None),
 (22, "Referral fee", "=-IF(C9=1,C6*C8,C21*C8)", EUR, None),
 (23, "FBA fulfilment fee", "=-C10", EUR, None),
 (24, "Storage", "=-C11", EUR, None),
 (25, "Cost of returns", "=-C12*C13*C21", EUR, None),
 (26, "Net after Amazon (payout per unit)", "=SUM(C21:C25)", EUR, "What Amazon disburses per unit sold, before your own costs."),
 (27, "Landed cost per unit", "=-(C14+C15+(C14+C15)*C16+C17)", EUR, "Duty on CIF base: (EXW + freight) × duty rate."),
 (28, "Turnover tax (on net revenue)", "=-C18*C21", EUR, None),
 (29, "Contribution margin per unit", "=C26+C27+C28", EUR, None),
 (30, "Margin on net revenue", '=IFERROR(C29/C21,"")', PCT, "Target: 35%+ before advertising."),
 (31, "Breakeven ACOS (ad console basis)", '=IFERROR(C29/C6,"")', PCT, "The console measures sales incl. VAT → divide by the GROSS price."),
 (32, "Max landed cost at 35% margin", "=C26+C28-0.35*C21", EUR, None),
 (33, "Max EXW price to quote supplier", "=(C32-C17-C15*(1+C16))/(1+C16)", EUR, "Hand THIS number to the supplier as your ceiling."),
]
for r, label, f, fmt, n in calc3:
    bold = r in (26, 29)
    put(ws, f"B{r}", label, F_BOLD if bold else F_BASE)
    frm(ws, f"C{r}", f, fmt=fmt, font=F_BOLD if bold else F_FORMULA)
    if n: note(ws, f"E{r}", n)
comment(ws, "C31", "Above this ACOS you lose money on every ad-driven unit. The Amazon ads console reports sales INCLUDING VAT, so this divides margin by the gross price — most templates get this wrong by 19%.")
comment(ws, "C18", "Georgian small-business status taxes turnover at 1%. Base assumed = ex-VAT sales; confirm with your accountant.")
note(ws, "B35", "Import VAT (19%) is paid at customs on customs value + duty and is recoverable via your German VAT registration — a cash-flow item (see Sheet 8), not a unit cost.")
note(ws, "B36", "Cross-check the FBA fee against real dimensions: €4.55 is not a 12 cm box. H10 exports sometimes mislabel inches as cm.")

# =====================================================================
# 4 SELECTION
# =====================================================================
ws = wb.create_sheet("4 Selection"); ws.sheet_properties.tabColor = TAB["choose"]
widths(ws, {"A": 2, "B": 62, "C": 10, "D": 2, "E": 62}); backlink(ws)
put(ws, "B2", "Selection checklist. Any NO in block A is an automatic reject.", F_TITLE)
put(ws, "B4", "A. Compliance & IP exclusions — YES only if this does NOT apply.", F_SECTION, fill=FILL_HDR)
put(ws, "C4", "Passed?", F_SECTION, fill=FILL_HDR)
blockA = [
 (5,  "Not worn for body protection (EU PPE Regulation 2016/425)", "Helmets, harnesses, goggles, gloves, hi-vis. Cat II needs a notified body."),
 (6,  "No battery, plug, motor or LED", "Triggers LVD, EMC, RoHS, WEEE, BattG."),
 (7,  "Not intended for children under 14", "Toy Safety Directive 2009/48/EC."),
 (8,  "No contact with food, drink or skin", "LFGB, cosmetics regulation, migration tests."),
 (9,  "No medical, health or veterinary claims", "MDR, health-claims rules."),
 (10, "Nothing load-bearing, nothing that arrests a fall", "The PPE trap under another name."),
 (11, "No blocking IP: designs, patents, trademarks checked", "EUIPO eSearch, DPMA register, Google Patents. Design infringement is the #1 private-label account killer in DE."),
 (12, "EU Responsible Person appointed (GPSR)", "Mandatory for a non-EU seller on any consumer product."),
 (13, "Packaging EPR done: LUCID + PPWR authorised representative", "LUCID mandatory since 2022; PPWR AR duty applies since 12 Aug 2026."),
]
for r, label, n in blockA:
    put(ws, f"B{r}", label, F_BASE, wrap=True); inp(ws, f"C{r}", "YES"); note(ws, f"E{r}", n)
put(ws, "B15", "B. Demand and competition", F_SECTION, fill=FILL_HDR); put(ws, "C15", "Passed?", F_SECTION, fill=FILL_HDR)
blockB = [
 (16, "Page-1 revenue of €40,000+/month", "Xray: sum of top-10 organic results."),
 (17, "Top ASIN holds under 40% of page revenue", "Higher = one brand plus noise."),
 (18, "At least 2 page-1 listings launched in last 12 months", "Proves entry still works."),
 (19, "Zero to two page-1 ASINs above 500 reviews", "Count REAL reviews — run Sheet 11's fake check first."),
 (20, "No Amazon Basics / Amazon Retail on page 1", ""),
 (21, "China + Hong Kong sellers under 60% of page 1", ""),
 (22, "A dominant negative-review theme of at least 40%", "From Sheet 11."),
 (23, "Weakest page-1 ASINs rated 4.3 or below", "A 4.7 field gives no point of attack."),
]
for r, label, n in blockB:
    put(ws, f"B{r}", label, F_BASE, wrap=True); inp(ws, f"C{r}", "YES")
    if n: note(ws, f"E{r}", n)
put(ws, "B25", "C. Seasonality and cash", F_SECTION, fill=FILL_HDR); put(ws, "C25", "Passed?", F_SECTION, fill=FILL_HDR)
blockC = [
 (26, "Amplitude below 2.5x, OR the order lands before the ramp", "Sheets 1 + 2."),
 (27, "Top-4 months hold under 60% of the year", ""),
 (28, "The order-by date is still in the future", "Sheet 2."),
 (29, "Contribution margin 35%+ before advertising", "Sheet 3."),
 (30, "Order + compliance + launch ads fit budget incl. reorder reserve", "Sheets 6 + 7 + 8."),
]
for r, label, n in blockC:
    put(ws, f"B{r}", label, F_BASE, wrap=True); inp(ws, f"C{r}", "YES")
    if n: note(ws, f"E{r}", n)
put(ws, "B32", "Block A result", F_BOLD); frm(ws, "C32", '=IF(COUNTIF(C5:C13,"YES")=9,"CLEAN","REJECT")', font=F_BOLD)
put(ws, "B33", "Block B score", F_BOLD); frm(ws, "C33", '=COUNTIF(C16:C23,"YES")&" / 8"', font=F_BOLD)
put(ws, "B34", "Block C score", F_BOLD); frm(ws, "C34", '=COUNTIF(C26:C30,"YES")&" / 5"', font=F_BOLD)
note(ws, "B36", "Go ahead only if A = CLEAN, B ≥ 7/8, C = 5/5.")

# =====================================================================
# 5 NICHE SCORECARD
# =====================================================================
ws = wb.create_sheet("5 Niche scorecard"); ws.sheet_properties.tabColor = TAB["choose"]
w = {"A": 2, "B": 48, "C": 14}
for j in range(8): w[get_column_letter(4 + j)] = 14
widths(ws, w); backlink(ws)
put(ws, "B2", "Niche comparison scorecard — up to 8 candidates", F_TITLE)
put(ws, "B3", "Fill the yellow inputs from Xray/Cerebro and from Sheets 1–3 run per niche. Any hard gate failure disqualifies regardless of score.", F_BASE, wrap=True)
put(ws, "B4", "Budget incl. reorder reserve (EUR)", F_BASE); inp(ws, "C4", 6500, fmt=EUR)
put(ws, "B5", "Inputs", F_SECTION, fill=FILL_HDR); put(ws, "C5", "Source", F_SECTION, fill=FILL_HDR)
for j in range(8): put(ws, f"{get_column_letter(4+j)}5", f"Niche {j+1}", F_SECTION, fill=FILL_HDR)
in5 = [
 (6,  "Niche / keyword", "—", None), (7,  "Avg monthly search volume (context, not scored)", "Cerebro", NUM),
 (8,  "Page-1 revenue, €/month", "Xray", NUM), (9,  "Top ASIN share of page-1 revenue", "Xray", PCT),
 (10, "Page-1 ASINs with >500 REAL reviews (count)", "Xray + Sheet 11", "0"),
 (11, "Page-1 listings launched in last 12 months", "Xray", "0"),
 (12, "Weak competitor rated ≤4.3 exists? (1/0)", "Xray", "0"),
 (13, "China + HK share of page 1", "Xray", PCT), (14, "Seasonality amplitude", "Sheet 1", IDX),
 (15, "Top-4 months' share of year", "Sheet 1", PCT), (16, "Contribution margin, % of net", "Sheet 3", PCT),
 (17, "First order + compliance cash need, €", "Sheets 6+8", NUM), (18, "Compliance block A passed? (YES/NO)", "Sheet 4", None),
]
for r, label, src, fmt in in5:
    put(ws, f"B{r}", label, F_BASE, wrap=True); note(ws, f"C{r}", src)
    for j in range(8): inp(ws, f"{get_column_letter(4+j)}{r}", None, fmt=fmt)
example = {6: "fahrradhelm", 7: 23500, 8: 95000, 9: 0.28, 10: 5, 11: 2, 12: 1,
           13: 0.55, 14: 5.15, 15: 0.54, 16: 0.40, 17: 9000, 18: "NO"}
for r, v in example.items(): ws[f"D{r}"].value = v
put(ws, "B20", "Scores (auto, 0–2 each)", F_SECTION, fill=FILL_HDR); put(ws, "C20", "Weight", F_SECTION, fill=FILL_HDR)
for j in range(8): put(ws, f"{get_column_letter(4+j)}20", "", F_SECTION, fill=FILL_HDR)
score_rows = [
 (21, "Demand: page-1 revenue", 0.15, '=IF({c}8="","",IF({c}8>=40000,2,IF({c}8>=20000,1,0)))'),
 (22, "Fragmentation: top-ASIN share", 0.10, '=IF({c}9="","",IF({c}9<0.3,2,IF({c}9<=0.4,1,0)))'),
 (23, "Review moat", 0.10, '=IF({c}10="","",IF({c}10<=2,2,IF({c}10<=4,1,0)))'),
 (24, "Fresh entrants", 0.10, '=IF({c}11="","",IF({c}11>=2,2,IF({c}11>=1,1,0)))'),
 (25, "Weak competitor to attack", 0.10, '=IF({c}12="","",IF({c}12=1,2,0))'),
 (26, "CN/HK saturation", 0.05, '=IF({c}13="","",IF({c}13<0.4,2,IF({c}13<0.6,1,0)))'),
 (27, "Seasonality amplitude", 0.15, '=IF({c}14="","",IF({c}14<1.5,2,IF({c}14<2.5,1,0)))'),
 (28, "Top-4 concentration", 0.05, '=IF({c}15="","",IF({c}15<0.45,2,IF({c}15<0.6,1,0)))'),
 (29, "Margin", 0.15, '=IF({c}16="","",IF({c}16>=0.35,2,IF({c}16>=0.3,1,0)))'),
 (30, "Cash fit vs budget", 0.05, '=IF({c}17="","",IF({c}17<=0.8*$C$4,2,IF({c}17<=$C$4,1,0)))'),
]
for r, label, wgt, f in score_rows:
    put(ws, f"B{r}", label, F_BASE); inp(ws, f"C{r}", wgt, fmt="0%")
    for j in range(8):
        col = get_column_letter(4 + j)
        frm(ws, f"{col}{r}", f.format(c=col), fmt="0")
put(ws, "B32", "Weighted score (0–100)", F_BOLD)
put(ws, "B33", "Hard gates", F_BOLD); put(ws, "B34", "Verdict", F_BOLD)
for j in range(8):
    col = get_column_letter(4 + j)
    frm(ws, f"{col}32", f'=IF(COUNT({col}21:{col}30)<10,"",ROUND(SUMPRODUCT($C$21:$C$30,{col}21:{col}30)/2*100,0))', fmt="0", font=F_BOLD)
    frm(ws, f"{col}33", f'=IF({col}18="NO","DISQUALIFIED: compliance",IF({col}14="","—",IF({col}14>=4,"DISQUALIFIED: single-season","OK")))', font=F_BOLD)
    frm(ws, f"{col}34", f'=IF({col}32="","",IF({col}33<>"OK","DROP",IF({col}32>=70,"INVESTIGATE",IF({col}32>=50,"PARK","DROP"))))', font=F_BOLD)
note(ws, "B36", "≥70 investigate · 50–69 park · <50 drop · any DISQUALIFIED = drop regardless of score.")
note(ws, "B37", "The score ranks research time. The final go/no-go is Sheets 1–4, never this score. Example column: a decent score killed by a hard gate — helmets are PPE Cat II.")

# =====================================================================
# 6 ORDER SIZE (newsvendor)
# =====================================================================
ws = wb.create_sheet("6 Order size"); ws.sheet_properties.tabColor = TAB["launch"]
widths(ws, {"A": 2, "B": 46, "C": 14, "D": 2, "E": 64}); backlink(ws)
put(ws, "B2", "First order size — newsvendor model", F_TITLE)
put(ws, "B3", "Balances the margin you lose on a missed sale against the cost of a leftover unit. Give it a pessimistic and an optimistic season estimate; it returns the profit-maximizing quantity.", F_BASE, wrap=True)
put(ws, "B5", "Inputs", F_SECTION, fill=FILL_HDR); put(ws, "C5", "", F_SECTION, fill=FILL_HDR)
in6 = [
 (6,  "Pessimistic season sales, units (P10)", 600, NUM, "Only ~10% chance sales come in below this. From Xray page-1 revenue × the share you think you can take."),
 (7,  "Optimistic season sales, units (P90)", 1400, NUM, "Only ~10% chance above this."),
 (8,  "Value recovered per leftover unit (EUR)", 5.0, EUR, "Clearance, Outlet, or removal — net of fees."),
 (9,  "Extra storage/removal per leftover unit (EUR)", 1.0, EUR, "Incl. aged-inventory surcharge risk."),
 (10, "MOQ (units)", 300, NUM, "Supplier minimum."),
 (11, "Budget for this PO (EUR)", 5000, EUR, "Cash available for goods (not ads)."),
]
for r, label, v, fmt, n in in6:
    put(ws, f"B{r}", label, F_BASE); inp(ws, f"C{r}", v, fmt=fmt); note(ws, f"E{r}", n)
put(ws, "B13", "From other sheets", F_SECTION, fill=FILL_HDR); put(ws, "C13", "", F_SECTION, fill=FILL_HDR)
put(ws, "B14", "Contribution margin per unit (Cu)", F_BASE); frm(ws, "C14", "='3 Unit economics'!C29", fmt=EUR, font=F_LINK)
put(ws, "B15", "Landed cost per unit", F_BASE); frm(ws, "C15", "=-'3 Unit economics'!C27", fmt=EUR, font=F_LINK)
put(ws, "B17", "Calculation", F_SECTION, fill=FILL_HDR); put(ws, "C17", "", F_SECTION, fill=FILL_HDR)
calc6 = [
 (18, "Expected season demand (mu)", "=(C6+C7)/2", NUM, None),
 (19, "Demand uncertainty (sigma)", "=(C7-C6)/2.5631", NUM, "Converts your P10–P90 range into a standard deviation."),
 (20, "Cost of one leftover unit (Co)", "=C15+C9-C8", EUR, "Landed + extra storage − recovery."),
 (21, "Critical ratio Cu/(Cu+Co)", '=IFERROR(C14/(C14+C20),"")', PCT, "The service level to buy. Above 50% = order MORE than average demand."),
 (22, "Optimal quantity (unconstrained)", '=IFERROR(ROUND(NORMINV(C21,C18,C19),0),"")', NUM, None),
 (23, "Max affordable units (budget / landed)", '=IFERROR(INT(C11/C15),"")', NUM, None),
 (24, "RECOMMENDED ORDER", '=IFERROR(MIN(MAX(C10,C22),C23),"")', NUM, "Optimum clamped to MOQ and budget."),
 (25, "z at recommended quantity", '=IFERROR((C24-C18)/C19,"")', IDX, None),
 (26, "Expected lost sales (units)", '=IFERROR(ROUND(C19*(NORMDIST(C25,0,1,FALSE)-C25*(1-NORMSDIST(C25))),0),"")', NUM, None),
 (27, "Expected leftover (units)", '=IFERROR(ROUND(C24-C18+C26,0),"")', NUM, None),
 (28, "Expected season contribution (EUR)", '=IFERROR(ROUND(C14*(C18-C26)-C20*C27,0),"")', NUM, None),
]
for r, label, f, fmt, n in calc6:
    bold = r in (21, 24)
    put(ws, f"B{r}", label, F_BOLD if bold else F_BASE)
    frm(ws, f"C{r}", f, fmt=fmt, font=F_BOLD if bold else F_FORMULA)
    if n: note(ws, f"E{r}", n)
comment(ws, "C21", "Why order above average when margin is high: a missed sale costs you the full margin; a leftover unit only costs landed minus recovery. When Cu > Co, under-ordering is the bigger mistake.")
frm(ws, "B30", '=IF(OR(C22="",C10=""),"",IF(C10>C22*1.2,"⚠ MOQ is far above the optimum — negotiate MOQ down or reconsider the product.",""))', font=F_BOLD)
frm(ws, "B31", '=IF(OR(C22="",C23=""),"",IF(C23<C22*0.8,"⚠ Budget caps you well below the optimum — plan the reorder date now (Sheet 8).",""))', font=F_BOLD)
note(ws, "B33", "For a year-round product this sizes the seasonal chunk; you can reorder, so lean toward the lower end and reorder fast.")

# =====================================================================
# 7 PPC PLANNER
# =====================================================================
ws = wb.create_sheet("7 PPC planner"); ws.sheet_properties.tabColor = TAB["launch"]
widths(ws, {"A": 2, "B": 40, "C": 12, "D": 11, "E": 9, "F": 11, "G": 24}); backlink(ws)
put(ws, "B2", "PPC planner — economics, launch budget, waste finder", F_TITLE)
put(ws, "B3", "Top: what you can afford to pay for ads. Bottom: paste your Search Term Report and see the money leaking.", F_BASE, wrap=True)
put(ws, "B5", "Ad economics", F_SECTION, fill=FILL_HDR); put(ws, "C5", "", F_SECTION, fill=FILL_HDR)
put(ws, "B6", "Gross price (incl. VAT)", F_BASE); frm(ws, "C6", "='3 Unit economics'!C6", fmt=EUR, font=F_LINK)
put(ws, "B7", "Contribution margin per unit", F_BASE); frm(ws, "C7", "='3 Unit economics'!C29", fmt=EUR, font=F_LINK)
put(ws, "B8", "Breakeven ACOS", F_BASE); frm(ws, "C8", '=IFERROR(C7/C6,"")', fmt=PCT, font=F_BOLD)
comment(ws, "C8", "ACOS = ad spend ÷ ad-attributed sales (console measures sales incl. VAT). At this ACOS an ad-driven unit earns exactly zero. Spend above it only deliberately, during launch.")
put(ws, "B9", "Breakeven ROAS", F_BASE); frm(ws, "C9", '=IFERROR(1/C8,"")', fmt=IDX)
put(ws, "B10", "Share of margin given to ads at steady state", F_BASE); inp(ws, "C10", 0.30, fmt=PCT)
note(ws, "E10", "0.30 = you keep 70% of margin, ads take 30%.")
put(ws, "B11", "Target ACOS (profit phase)", F_BOLD); frm(ws, "C11", '=IFERROR(C8*C10,"")', fmt=PCT, font=F_BOLD)
put(ws, "B12", "Target ROAS", F_BASE); frm(ws, "C12", '=IFERROR(1/C11,"")', fmt=IDX)
put(ws, "B14", "Launch budget", F_SECTION, fill=FILL_HDR); put(ws, "C14", "", F_SECTION, fill=FILL_HDR)
put(ws, "B15", "Target ad-driven orders per day", F_BASE); inp(ws, "C15", 3, fmt="0")
put(ws, "B16", "Estimated CPC (EUR)", F_BASE); inp(ws, "C16", 0.80, fmt=EUR)
put(ws, "B17", "Estimated conversion rate", F_BASE); inp(ws, "C17", 0.10, fmt=PCT)
note(ws, "E17", "Clicks per order = 1 / CVR. 10% is a decent listing; below 7% fix the listing, not the bids.")
put(ws, "B18", "Daily ad spend", F_BASE); frm(ws, "C18", '=IFERROR(C15*C16/C17,"")', fmt=EUR)
put(ws, "B19", "Launch length (weeks)", F_BASE); inp(ws, "C19", 4, fmt="0")
put(ws, "B20", "Launch ad budget", F_BOLD); frm(ws, "C20", '=IFERROR(C18*7*C19,"")', fmt=EUR, font=F_BOLD)
put(ws, "B21", "TACOS = ad spend ÷ TOTAL sales (ads + organic). Track it monthly in Sheet 9 — healthy: falling while sales grow.", F_NOTE, wrap=True)
put(ws, "B22", "Waste finder — summary", F_SECTION, fill=FILL_HDR); put(ws, "C22", "", F_SECTION, fill=FILL_HDR)
put(ws, "B23", "Flag 0-order terms with clicks ≥", F_BASE); inp(ws, "C23", 15, fmt="0")
put(ws, "B24", "…or spend ≥ (EUR)", F_BASE); inp(ws, "C24", 10, fmt=EUR)
put(ws, "B25", "Wasted spend on 0-order terms", F_BOLD)
frm(ws, "C25", '=IFERROR(SUMIFS($D$30:$D$229,$E$30:$E$229,0,$B$30:$B$229,"<>"),0)', fmt=EUR, font=F_BOLD)
put(ws, "B26", "Negative-keyword candidates", F_BOLD)
frm(ws, "C26", '=COUNTIF($G$30:$G$229,"ADD AS NEGATIVE")', fmt="0", font=F_BOLD)
put(ws, "B28", "Paste your Search Term Report below (Ads console → Sponsored Products → Reports → Search term, 30–60 days). First 3 rows are examples — overwrite them.", F_NOTE, wrap=True)
for col, h in [("B","Search term"),("C","Clicks"),("D","Spend €"),("E","Orders"),("F","Sales €"),("G","Flag")]:
    put(ws, f"{col}29", h, F_SECTION, fill=FILL_HDR)
examples7 = [("fahrradhelm herren", 42, 31.50, 3, 134.85), ("helm kinder", 23, 18.20, 0, 0), ("motorradhelm", 35, 26.80, 0, 0)]
for r in range(30, 230):
    if r - 30 < len(examples7):
        t, cl, sp, o, sa = examples7[r - 30]
        inp(ws, f"B{r}", t); inp(ws, f"C{r}", cl, fmt="0"); inp(ws, f"D{r}", sp, fmt=EUR)
        inp(ws, f"E{r}", o, fmt="0"); inp(ws, f"F{r}", sa, fmt=EUR)
    else:
        for col, fmt in [("B",None),("C","0"),("D",EUR),("E","0"),("F",EUR)]:
            inp(ws, f"{col}{r}", None, fmt=fmt)
    frm(ws, f"G{r}", f'=IF(COUNT(C{r}:F{r})<4,"",IF(AND(E{r}=0,OR(C{r}>=$C$23,D{r}>=$C$24)),"ADD AS NEGATIVE",IF(AND(F{r}>0,D{r}/F{r}>$C$8),"ACOS > BREAKEVEN","OK")))')
ws.freeze_panes = "A30"

# =====================================================================
# 8 CASH FLOW
# =====================================================================
ws = wb.create_sheet("8 Cash flow"); ws.sheet_properties.tabColor = TAB["launch"]
widths(ws, {"A": 2, "B": 9, "C": 13, "D": 13, "E": 13, "F": 13, "G": 13, "H": 14, "I": 40}); backlink(ws)
put(ws, "B2", "First-order cash plan — 12 months", F_TITLE)
put(ws, "B3", "Money leaves in lumps and comes back as a drip. If the cumulative line goes below zero, the plan is not funded.", F_BASE, wrap=True)
put(ws, "B5", "Inputs", F_SECTION, fill=FILL_HDR); put(ws, "C5", "", F_SECTION, fill=FILL_HDR)
put(ws, "B6", "Order quantity", F_BASE); frm(ws, "C6", "='6 Order size'!C24", fmt=NUM, font=F_LINK)
put(ws, "B7", "Landed cost per unit", F_BASE); frm(ws, "C7", "='6 Order size'!C15", fmt=EUR, font=F_LINK)
put(ws, "B8", "PO value (goods, landed)", F_BASE); frm(ws, "C8", '=IFERROR(C6*C7,"")', fmt=EUR)
put(ws, "B9", "Deposit share at order (month 0)", F_BASE); inp(ws, "C9", 0.30, fmt=PCT)
put(ws, "B10", "Balance paid in month #", F_BASE); inp(ws, "C10", 2, fmt="0")
put(ws, "B11", "Goods arrive in month #", F_BASE); inp(ws, "C11", 3, fmt="0")
put(ws, "B12", "Import VAT paid at arrival (≈19% of PO)", F_BASE); frm(ws, "C12", '=IFERROR(ROUND(0.19*C8,0),"")', fmt=EUR)
comment(ws, "C12", "Einfuhrumsatzsteuer: paid at customs, recovered via your German monthly VAT return ~2 months later. It is a LOAN to the state, not a cost — but the cash must exist.")
put(ws, "B13", "Import VAT refunded after (months)", F_BASE); inp(ws, "C13", 2, fmt="0")
put(ws, "B14", "Amazon payout per unit sold", F_BASE); frm(ws, "C14", "='3 Unit economics'!C26", fmt=EUR, font=F_LINK)
put(ws, "B15", "Starting cash available (EUR)", F_BASE); inp(ws, "C15", 6500, fmt=EUR)
row = 16
put(ws, f"B{row}", "Month", F_SECTION, fill=FILL_HDR)
put(ws, f"C{row}", "Units sold", F_SECTION, fill=FILL_HDR)
put(ws, f"D{row}", "Ads+fixed €", F_SECTION, fill=FILL_HDR)
put(ws, f"E{row}", "Cash out €", F_SECTION, fill=FILL_HDR)
put(ws, f"F{row}", "Cash in €", F_SECTION, fill=FILL_HDR)
put(ws, f"G{row}", "Net €", F_SECTION, fill=FILL_HDR)
put(ws, f"H{row}", "Cumulative €", F_SECTION, fill=FILL_HDR)
units_plan = [0,0,0,60,90,110,120,120,100,80,60,40,20]
ads_plan   = [0,0,0,700,700,500,400,400,350,300,250,200,150]
for i in range(13):
    r = 17 + i
    put(ws, f"B{r}", i, F_BASE, fmt="0")
    inp(ws, f"C{r}", units_plan[i], fmt="0"); inp(ws, f"D{r}", ads_plan[i], fmt=NUM)
    frm(ws, f"E{r}", f'=ROUND(IF($B{r}=0,$C$9*$C$8,0)+IF($B{r}=$C$10,(1-$C$9)*$C$8,0)+IF($B{r}=$C$11,$C$12,0)+D{r},0)', fmt=NUM)
    frm(ws, f"F{r}", f'=ROUND(C{r}*$C$14+IF($B{r}=$C$11+$C$13,$C$12,0),0)', fmt=NUM)
    frm(ws, f"G{r}", f'=F{r}-E{r}', fmt=NUM)
    frm(ws, f"H{r}", f'=$C$15+G{r}' if i == 0 else f'=H{r-1}+G{r}', fmt=NUM, font=F_BOLD)
note(ws, "I17", "Units plan: copy the seasonal shape from Sheet 1; example follows a spring ramp.")
note(ws, "I20", "Ads column: launch-heavy, from Sheet 7's launch budget, then steady state.")
put(ws, "B32", "Lowest cash point", F_BOLD); frm(ws, "C32", "=MIN(H17:H29)", fmt=NUM, font=F_BOLD)
frm(ws, "B33", '=IF(C32<0,"⚠ Cash goes below zero — cut the order (Sheet 6), stretch supplier terms, or add capital before ordering.","OK — the plan stays funded.")', font=F_BOLD)
note(ws, "B35", "Not included: your salary, Georgian 1% tax timing, the reorder. Add the reorder as extra cash-out rows when you plan it — running out of stock in season kills launches (Sheet 4, block C).")

# =====================================================================
# 9 MONTHLY P&L
# =====================================================================
ws = wb.create_sheet("9 Monthly P&L"); ws.sheet_properties.tabColor = TAB["operate"]
widths(ws, {"A": 2, "B": 8, "C": 8, "D": 12, "E": 12, "F": 11, "G": 10, "H": 13, "I": 11, "J": 11, "K": 12, "L": 10, "M": 9}); backlink(ws)
put(ws, "B2", "Monthly P&L — the real profit, per month", F_TITLE)
put(ws, "B3", "Fill monthly from Payments → Reports repository → Date Range Report (Summary) and the Ads console. Enter costs as POSITIVE numbers. Row 6 is an example — overwrite it.", F_BASE, wrap=True)
hdrs9 = ["Month","Units","Gross sales €","Amazon fees €","Refunds €","PPC €","Storage+other €","COGS € (auto)","Tax € (auto)","Net profit €","Margin %","TACOS"]
for j, h in enumerate(hdrs9): put(ws, f"{get_column_letter(2+j)}5", h, F_SECTION, fill=FILL_HDR)
for i, m in enumerate(months):
    r = 6 + i
    put(ws, f"B{r}", m, F_BASE)
    for col in ["C","D","E","F","G","H"]: inp(ws, f"{col}{r}", None, fmt=NUM)
    frm(ws, f"I{r}", f"=IF(C{r}=\"\",\"\",ROUND(C{r}*'6 Order size'!$C$15,0))", fmt=NUM)
    frm(ws, f"J{r}", f"=IF(D{r}=\"\",\"\",ROUND('3 Unit economics'!$C$18*D{r}/(1+'3 Unit economics'!$C$7),0))", fmt=NUM)
    frm(ws, f"K{r}", f"=IF(D{r}=\"\",\"\",ROUND(D{r}/(1+'3 Unit economics'!$C$7)-E{r}-F{r}-G{r}-H{r}-I{r}-J{r},0))", fmt=NUM, font=F_BOLD)
    frm(ws, f"L{r}", f"=IF(D{r}=\"\",\"\",K{r}/(D{r}/(1+'3 Unit economics'!$C$7)))", fmt=PCT)
    frm(ws, f"M{r}", f'=IF(OR(D{r}="",G{r}=""),"",G{r}/D{r})', fmt=PCT)
for c, v in {"C6":90, "D6":5399, "E6":1260, "F6":180, "G6":620, "H6":90}.items():
    ws[c].value = v
put(ws, "B18", "Total", F_BOLD)
for col in ["C","D","E","F","G","H","I","J","K"]:
    frm(ws, f"{col}18", f"=SUM({col}6:{col}17)", fmt=NUM, font=F_BOLD)
frm(ws, "L18", "=IF(D18=0,\"\",K18/(D18/(1+'3 Unit economics'!$C$7)))", fmt=PCT, font=F_BOLD)
frm(ws, "M18", '=IF(D18=0,"",G18/D18)', fmt=PCT, font=F_BOLD)
comment(ws, "M5", "TACOS = ad spend ÷ TOTAL gross sales. Falling TACOS while sales grow = organic rank is compounding. Rising TACOS = you are buying every sale.")
note(ws, "B20", "Net profit = gross/1.19 − fees − refunds − PPC − storage − COGS − 1% turnover tax. VAT never reaches you (deemed supplier), so it is excluded on both sides.")
note(ws, "B21", "Cross-check: Margin % here should converge to Sheet 3's margin as volumes stabilize. If it is far lower, hunt the difference: storage, refunds, PPC.")
ws.freeze_panes = "A6"

# =====================================================================
# 10 MONEY RECOVERY
# =====================================================================
ws = wb.create_sheet("10 Money recovery"); ws.sheet_properties.tabColor = TAB["operate"]
widths(ws, {"A": 2, "B": 14, "C": 16, "D": 14, "E": 9, "F": 30, "G": 40}); backlink(ws)
put(ws, "B2", "Money recovery — what Amazon owes you", F_TITLE)
put(ws, "B3", "Sellers lose 1–3% of revenue to unclaimed reimbursements. The claim window is 60 DAYS, and since Mar 2025 auto-reimbursements cover only your manufacturing cost — check monthly, claim actively.", F_BASE, wrap=True)
put(ws, "B5", "Adjustments without a matching reimbursement:", F_BOLD)
frm(ws, "C5", '=COUNTIF($F$12:$F$61,"NOT REIMBURSED — check & claim")', fmt="0", font=F_BOLD)
comment(ws, "C5", "Each flag = a lost/damaged/disposed event with no reimbursement row for the same FNSKU. Open a case: Help → FBA → Something else → Inventory reimbursement.")
put(ws, "B7", "A — paste Inventory Adjustments (Reports → Fulfilment by Amazon → Inventory Adjustments; filter Lost, Damaged, Disposed). Rows 12–13 are examples.", F_NOTE, wrap=True)
for col, h in [("B","Date"),("C","FNSKU"),("D","Reason"),("E","Qty"),("F","Status")]:
    put(ws, f"{col}11", h, F_SECTION, fill=FILL_HDR)
exA = [(datetime.date(2026,7,3), "X001ABC", "Lost", 2), (datetime.date(2026,7,19), "X002DEF", "Damaged", 1)]
for r in range(12, 62):
    if r - 12 < len(exA):
        d, f_, rs, q = exA[r - 12]
        inp(ws, f"B{r}", d, fmt=DATEF); inp(ws, f"C{r}", f_); inp(ws, f"D{r}", rs); inp(ws, f"E{r}", q, fmt="0")
    else:
        inp(ws, f"B{r}", None, fmt=DATEF); inp(ws, f"C{r}", None); inp(ws, f"D{r}", None); inp(ws, f"E{r}", None, fmt="0")
    frm(ws, f"F{r}", f'=IF($C{r}="","",IF(COUNTIF($C$68:$C$117,$C{r})>0,"Reimbursement found — verify amount","NOT REIMBURSED — check & claim"))')
put(ws, "B64", "B — paste Reimbursements (Reports → Fulfilment by Amazon → Payments → Reimbursements). Row 68 is an example.", F_NOTE, wrap=True)
for col, h in [("B","Date"),("C","FNSKU"),("D","Amount €")]:
    put(ws, f"{col}67", h, F_SECTION, fill=FILL_HDR)
inp(ws, "B68", datetime.date(2026,7,21), fmt=DATEF); inp(ws, "C68", "X001ABC"); inp(ws, "D68", 18.40, fmt=EUR)
for r in range(69, 118):
    inp(ws, f"B{r}", None, fmt=DATEF); inp(ws, f"C{r}", None); inp(ws, f"D{r}", None, fmt=EUR)
put(ws, "B120", "C — FBA fee sanity check (fee overcharges from wrong measurements)", F_SECTION, fill=FILL_HDR)
put(ws, "B121", "Length (cm)", F_BASE);  inp(ws, "C121", 25.0, fmt="0.0")
put(ws, "B122", "Width (cm)", F_BASE);   inp(ws, "C122", 18.0, fmt="0.0")
put(ws, "B123", "Height (cm)", F_BASE);  inp(ws, "C123", 4.0, fmt="0.0")
put(ws, "B124", "Unit weight (g)", F_BASE); inp(ws, "C124", 350, fmt="0")
put(ws, "B125", "FBA fee you are charged (EUR)", F_BASE); inp(ws, "C125", 4.55, fmt=EUR)
put(ws, "B126", "Volume (litres)", F_BASE); frm(ws, "C126", "=ROUND(C121*C122*C123/1000,2)", fmt="0.00")
put(ws, "B127", "Longest side (cm)", F_BASE); frm(ws, "C127", "=MAX(C121:C123)", fmt="0.0")
note(ws, "B129", "Compare volume, longest side and weight against the current size tiers in Seller Central → 'FBA fulfilment fees rate card'. If your true tier is smaller than the fee implies, request a re-measure (case: FBA → Fee dispute). Amazon's measurements drift — recheck after every restock.")
put(ws, "B131", "Monthly audit routine", F_SECTION, fill=FILL_HDR); put(ws, "C131", "Done?", F_SECTION, fill=FILL_HDR)
audit = ["Adjustments report pulled and pasted (last 60 days)", "Reimbursements report pulled and pasted",
         "All flags above opened as cases", "Returns: refunds older than 60 days without item back → claim",
         "FBA fee vs dimensions rechecked after last restock"]
for i, a in enumerate(audit):
    r = 132 + i
    put(ws, f"B{r}", a, F_BASE, wrap=True); inp(ws, f"C{r}", "NO")

# =====================================================================
# 11 REVIEW THEMES
# =====================================================================
ws = wb.create_sheet("11 Review themes"); ws.sheet_properties.tabColor = TAB["operate"]
widths(ws, {"A": 2, "B": 42, "C": 12, "D": 10, "E": 58}); backlink(ws)
put(ws, "B2", "Review themes — find the complaint to attack", F_TITLE)
put(ws, "B3", "Export 1–3★ reviews of the top 5–10 ASINs (H10 Review Insights, last 12 months). Ask an AI assistant: 'cluster these reviews into complaint themes with counts'. Enter the themes below. Example data shown — overwrite.", F_BASE, wrap=True)
put(ws, "B5", "Niche / ASIN set", F_BASE); inp(ws, "C5", "satteltasche fahrrad")
for col, h in [("B","Complaint theme"),("C","Count"),("D","Share")]:
    put(ws, f"{col}7", h, F_SECTION, fill=FILL_HDR)
ex11 = [("Wobbles / mount comes loose", 46), ("Straps tear", 21), ("Not actually waterproof", 18),
        ("Too small / fit unclear", 15), ("Zipper fails", 8)]
for r in range(8, 18):
    if r - 8 < len(ex11):
        t, c = ex11[r - 8]; inp(ws, f"B{r}", t); inp(ws, f"C{r}", c, fmt="0")
    else:
        inp(ws, f"B{r}", None); inp(ws, f"C{r}", None, fmt="0")
    frm(ws, f"D{r}", f'=IF(C{r}="","",C{r}/SUM($C$8:$C$17))', fmt=PCT)
put(ws, "B19", "Dominant theme", F_BOLD)
frm(ws, "C19", '=IFERROR(INDEX(B8:B17,MATCH(MAX(C8:C17),C8:C17,0)),"—")', font=F_BOLD)
put(ws, "B20", "Its share of all complaints", F_BOLD)
frm(ws, "C20", '=IFERROR(MAX(C8:C17)/SUM(C8:C17),"")', fmt=PCT, font=F_BOLD)
comment(ws, "C20", "≥40% on a CORE feature = a product gap worth building against (Sheet 4 row). A dominant theme about shipping/packaging is Amazon's problem, not a product gap.")
put(ws, "B21", "Verdict", F_BOLD)
frm(ws, "C21", '=IF(C20="","",IF(C20>=0.4,"≥40% — you have an attack hypothesis (Sheet 4, block B).","No dominant theme — weak differentiation basis."))', font=F_BOLD)
put(ws, "B23", "Fake-review check — YES = red flag present", F_SECTION, fill=FILL_HDR); put(ws, "C23", "", F_SECTION, fill=FILL_HDR)
flags = [
 "Review bursts: many reviews landing within a few days",
 "Rating drifting down over 6 months after an early 4.7+",
 "Generic / repeated phrasing, or wrong product described",
 "High share of unverified purchases",
 "Oldest reviews describe a DIFFERENT product (listing merge)",
 "Reviewer profiles: many unrelated 5★ reviews same day",
]
for i, f_ in enumerate(flags):
    r = 24 + i
    put(ws, f"B{r}", f_, F_BASE, wrap=True); inp(ws, f"C{r}", "NO")
put(ws, "B30", "Red flags", F_BOLD); frm(ws, "C30", '=COUNTIF(C24:C29,"YES")', fmt="0", font=F_BOLD)
put(ws, "B31", "Verdict", F_BOLD)
frm(ws, "C31", '=IF(C30>=2,"Treat this review wall as partly fake — the moat is weaker than it looks (adjust Sheet 5, review moat).","No strong fake signals.")', font=F_BOLD)
note(ws, "B33", "Keepa's review-count history chart shows bursts in ten seconds per ASIN. Method: He, Hollenbeck & Proserpio, Marketing Science 2022.")

wb.save(OUT)
print("saved", OUT)
